#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test — TSETMC Active Endpoints (شناخته‌شده از تست قبلی)
=========================================================

این اسکریپت فقط همان 55+1 endpoint‌ای را که در تست‌های قبلی به‌عنوان "فعال با داده واقعی"
شناسایی شدند (فایل مرجع tsetmc_active_endpoints.md) دوباره تست می‌کند — برخلاف
iran_market_api_tester_v2.py که کل inventory (شامل مسیرهای مرده/شخص‌ثالث) را تست می‌کند.

به‌روزرسانی: یک endpoint جدید اضافه شد — "Market watch Excel (legacy)" روی
old.tsetmc.com که جایگزین واقعیِ GetMarketWatch (که همیشه خالی است) است.
این یکی فایل Excel برمی‌گرداند، نه JSON؛ برای شمارش ردیف‌هایش pip install
openpyxl لازم است (اختیاری — بدون آن هم بقیه‌ی endpointها normal کار می‌کنند).

هدف: تشخیص سریع اینکه آیا این مسیرهای شناخته‌شده هنوز زنده‌اند یا سرویس تغییر کرده.

کاربرد:
    pip install requests openpyxl
    python test_active_tsetmc_endpoints.py
    python test_active_tsetmc_endpoints.py --inscode 17914401175772326 --date 20260701
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import time
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from urllib.parse import urlparse

import requests
from io import BytesIO

try:
    import openpyxl
except Exception:
    openpyxl = None

# ---------------------------------------------------------------------------
# 55 endpoint فعالِ شناخته‌شده (استخراج‌شده از نتایج تست واقعی قبلی)
# ---------------------------------------------------------------------------
ENDPOINTS: List[Dict[str, str]] = [
    {"category": "ClosingPrice", "name": "ChartData daily", "template": "https://cdn.tsetmc.com/api/ClosingPrice/GetChartData/{InsCode}/D", "method": "GET", "description": "OHLCVT chart data, daily resolution"},
    {"category": "ClosingPrice", "name": "ClosingPrice daily", "template": "https://cdn.tsetmc.com/api/ClosingPrice/GetClosingPriceDaily/{InsCode}/{DEven}", "method": "GET", "description": "Single day closing price summary"},
    {"category": "ClosingPrice", "name": "ClosingPrice daily list top", "template": "https://cdn.tsetmc.com/api/ClosingPrice/GetClosingPriceDailyList/{InsCode}/{Top}", "method": "GET", "description": "Historical closing prices; Top=0 returns all"},
    {"category": "ClosingPrice", "name": "ClosingPrice daily list all", "template": "https://cdn.tsetmc.com/api/ClosingPrice/GetClosingPriceDailyList/{InsCode}/0", "method": "GET", "description": "All historical closing prices"},
    {"category": "ClosingPrice", "name": "ClosingPrice daily list CSV", "template": "https://cdn.tsetmc.com/api/ClosingPrice/GetClosingPriceDailyListCSV/{InsCode}/0", "method": "GET", "description": "All historical closing prices as CSV-like output"},
    {"category": "ClosingPrice", "name": "ClosingPrice history", "template": "https://cdn.tsetmc.com/api/ClosingPrice/GetClosingPriceHistory/{InsCode}/{DEven}", "method": "GET", "description": "Old ClosingPriceData intraday/history snapshots"},
    {"category": "ClosingPrice", "name": "ClosingPrice info today", "template": "https://cdn.tsetmc.com/api/ClosingPrice/GetClosingPriceInfo/{InsCode}", "method": "GET", "description": "Today price summary and instrument state"},
    {"category": "ClosingPrice", "name": "Instruments history in day", "template": "https://cdn.tsetmc.com/api/ClosingPrice/GetInstrmentsHistoryInDay/{DEven}", "method": "GET", "description": "Price info of all instruments on a specific date"},
    {"category": "ClosingPrice", "name": "Instrument calendar", "template": "https://cdn.tsetmc.com/api/ClosingPrice/GetInstrumentCalendar/{InsCode}", "method": "GET", "description": "Instrument trading calendar with closing prices and volumes"},
    {"category": "ClosingPrice", "name": "Market watch", "template": "https://cdn.tsetmc.com/api/ClosingPrice/GetMarketWatch", "method": "GET", "description": "Market watch data; may return empty from CDN without params/session"},
    {"category": "ClosingPrice", "name": "Market watch Excel (legacy)", "template": "https://old.tsetmc.com/tsev2/excel/MarketWatchPlus.aspx?d=0", "method": "GET", "description": "Legacy full market watch as Excel; ~3100 rows, no InsCode/sector, but confirmed working alternative to the always-empty API version", "timeout": 60.0, "max_retries": 1},
    {"category": "ClosingPrice", "name": "Price adjust by flow", "template": "https://cdn.tsetmc.com/api/ClosingPrice/GetPriceAdjustByFlow/{Flow}/{Top}", "method": "GET", "description": "Price adjustment events by market flow"},
    {"category": "ClosingPrice", "name": "Price adjust list", "template": "https://cdn.tsetmc.com/api/ClosingPrice/GetPriceAdjustList/{InsCode}", "method": "GET", "description": "Price adjustment events for instrument"},
    {"category": "ClosingPrice", "name": "Related company", "template": "https://cdn.tsetmc.com/api/ClosingPrice/GetRelatedCompany/{CSecVal}", "method": "GET", "description": "Instruments and 30-day history in industry sector"},
    {"category": "BestLimits", "name": "BestLimits now", "template": "https://cdn.tsetmc.com/api/BestLimits/{InsCode}", "method": "GET", "description": "Current 5-level order book / best limits"},
    {"category": "BestLimits", "name": "BestLimits historical", "template": "https://cdn.tsetmc.com/api/BestLimits/{InsCode}/{DEven}", "method": "GET", "description": "Historical order book snapshots for one trading day"},
    {"category": "Trade", "name": "Trade now", "template": "https://cdn.tsetmc.com/api/Trade/GetTrade/{InsCode}", "method": "GET", "description": "Trades for last trading day"},
    {"category": "Trade", "name": "Trade history", "template": "https://cdn.tsetmc.com/api/Trade/GetTradeHistory/{InsCode}/{DEven}/false", "method": "GET", "description": "Historical intraday trades"},
    {"category": "Trade", "name": "Trade intraday", "template": "https://cdn.tsetmc.com/api/Trade/GetTradeIntraDay/{InsCode}", "method": "GET", "description": "Intraday aggregated trades for last trading day"},
    {"category": "Trade", "name": "Trade volume", "template": "https://cdn.tsetmc.com/api/Trade/GetTradeVolume/{InsCode}", "method": "GET", "description": "Trade volume distribution by price/time"},
    {"category": "ClientType", "name": "ClientType now", "template": "https://cdn.tsetmc.com/api/ClientType/GetClientType/{InsCode}/1/0", "method": "GET", "description": "Current real/legal buyer-seller data"},
    {"category": "ClientType", "name": "ClientType all", "template": "https://cdn.tsetmc.com/api/ClientType/GetClientTypeAll", "method": "GET", "description": "Client type data for all instruments from last trading day"},
    {"category": "ClientType", "name": "ClientType history", "template": "https://cdn.tsetmc.com/api/ClientType/GetClientTypeHistory/{InsCode}/{DEven}", "method": "GET", "description": "Historical real/legal buyer-seller data for instrument/day"},
    {"category": "Instrument", "name": "Instrument history", "template": "https://cdn.tsetmc.com/api/Instrument/GetInstrumentHistory/{InsCode}/{DEven}", "method": "GET", "description": "Historical simple instrument metadata for a day"},
    {"category": "Instrument", "name": "Instrument identity", "template": "https://cdn.tsetmc.com/api/Instrument/GetInstrumentIdentity/{InsCode}", "method": "GET", "description": "Instrument identity, sector, sub-sector, names"},
    {"category": "Instrument", "name": "Instrument info", "template": "https://cdn.tsetmc.com/api/Instrument/GetInstrumentInfo/{InsCode}", "method": "GET", "description": "Instrument full information, EPS, sector, static thresholds"},
    {"category": "Instrument", "name": "Instrument search", "template": "https://cdn.tsetmc.com/api/Instrument/GetInstrumentSearch/{Query}", "method": "GET", "description": "Instrument search by keyword"},
    {"category": "Instrument", "name": "Share change", "template": "https://cdn.tsetmc.com/api/Instrument/GetInstrumentShareChange/{InsCode}", "method": "GET", "description": "Share changes for instrument"},
    {"category": "Instrument", "name": "Share change by flow", "template": "https://cdn.tsetmc.com/api/Instrument/GetInstrumentShareChangeByFlow/{Flow}/{Top}", "method": "GET", "description": "Share changes by market flow"},
    {"category": "MarketData", "name": "Instrument state", "template": "https://cdn.tsetmc.com/api/MarketData/GetInstrumentState/{InsCode}/{DEven}", "method": "GET", "description": "Instrument state for day"},
    {"category": "MarketData", "name": "Instrument state top", "template": "https://cdn.tsetmc.com/api/MarketData/GetInstrumentStateTop/{Top}", "method": "GET", "description": "Recently changed instrument states"},
    {"category": "MarketData", "name": "Instrument statistic", "template": "https://cdn.tsetmc.com/api/MarketData/GetInstrumentStatistic/{InsCode}", "method": "GET", "description": "Instrument statistics such as average value/volume/trades"},
    {"category": "MarketData", "name": "Inst value all params", "template": "https://cdn.tsetmc.com/api/MarketData/GetInstValueAllInstAllParam", "method": "GET", "description": "All instrument value parameters for all instruments; very large"},
    {"category": "MarketData", "name": "Market overview", "template": "https://cdn.tsetmc.com/api/MarketData/GetMarketOverview/{Flow}", "method": "GET", "description": "Market overview by flow with index values and activity"},
    {"category": "MarketData", "name": "Sectors summary", "template": "https://cdn.tsetmc.com/api/MarketData/GetSectorsSummary", "method": "GET", "description": "Sectors summary"},
    {"category": "MarketData", "name": "Static threshold", "template": "https://cdn.tsetmc.com/api/MarketData/GetStaticThreshold/{InsCode}/{DEven}", "method": "GET", "description": "Static price thresholds for day"},
    {"category": "Fund", "name": "ETF by inscode", "template": "https://cdn.tsetmc.com/api/Fund/GetETFByInsCode/{InsCode}", "method": "GET", "description": "ETF/fund data including redemption/subscription NAV-like values"},
    {"category": "Fund", "name": "Fund detail 0", "template": "https://cdn.tsetmc.com/api/Fund/GetFundInDetail/0", "method": "GET", "description": "Fund details and NAV stats"},
    {"category": "Fund", "name": "Fund detail 1", "template": "https://cdn.tsetmc.com/api/Fund/GetFundInDetail/1", "method": "GET", "description": "Fund details variant"},
    {"category": "Shareholder", "name": "Shareholder changes old", "template": "https://cdn.tsetmc.com/api/Shareholder/{InsCode}/{DEven}", "method": "GET", "description": "Shareholder changes at day"},
    {"category": "Shareholder", "name": "Shareholder last", "template": "https://cdn.tsetmc.com/api/Shareholder/GetInstrumentShareHolderLast/{InsCode}", "method": "GET", "description": "Latest shareholders for instrument"},
    {"category": "Shareholder", "name": "Shareholder changes all", "template": "https://cdn.tsetmc.com/api/Shareholder/GetShareHolderChanges/false", "method": "GET", "description": "All shareholder changes in recent days"},
    {"category": "Shareholder", "name": "Shareholder company list", "template": "https://cdn.tsetmc.com/api/Shareholder/GetShareHolderCompanyList/{ShareHolderShareID}", "method": "GET", "description": "Other holdings of shareholder"},
    {"category": "Shareholder", "name": "Shareholder history", "template": "https://cdn.tsetmc.com/api/Shareholder/GetShareHolderHistory/{InsCode}/{DEven}/{Top}", "method": "GET", "description": "Shareholder changes for instrument/day"},
    {"category": "Codal", "name": "Prepared data top", "template": "https://cdn.tsetmc.com/api/Codal/GetPreparedData/{Top}", "method": "GET", "description": "Recent codal notifications"},
    {"category": "Codal", "name": "Prepared data by inscode", "template": "https://cdn.tsetmc.com/api/Codal/GetPreparedDataByInsCode/{Top}/{InsCode}", "method": "GET", "description": "Codal notifications for instrument"},
    {"category": "Codal", "name": "Codal publisher by symbol", "template": "https://cdn.tsetmc.com/api/Codal/GetCodalPublisherBySymbol/{Symbol}", "method": "GET", "description": "Codal publisher lookup by symbol"},
    {"category": "Codal", "name": "File attachment by row id", "template": "https://cdn.tsetmc.com/api/Codal/GetFileAttachmentByMainTableRowId/{MainTableRowId}", "method": "GET", "description": "Codal attachment by main table row id"},
    {"category": "Codal", "name": "File attachment by tracing", "template": "https://cdn.tsetmc.com/api/Codal/GetFileAttachmentByTracingNo/{TracingNo}", "method": "GET", "description": "Codal attachment by tracing no"},
    {"category": "Msg", "name": "Messages by flow", "template": "https://cdn.tsetmc.com/api/Msg/GetMsgByFlow/{Flow}/{Top}", "method": "GET", "description": "Supervisor messages by flow"},
    {"category": "Msg", "name": "Messages by inscode", "template": "https://cdn.tsetmc.com/api/Msg/GetMsgByInsCode/{InsCode}", "method": "GET", "description": "Supervisor messages for instrument"},
    {"category": "Energy", "name": "Energy auction by id", "template": "https://cdn.tsetmc.com/api/Energy/GetAuctionById/{AuctionId}", "method": "GET", "description": "Energy exchange auction by id"},
    {"category": "Energy", "name": "Energy auction trade by id", "template": "https://cdn.tsetmc.com/api/Energy/GetAuctionTradeById/{AuctionId}", "method": "GET", "description": "Energy auction trades by id"},
    {"category": "StaticData", "name": "Static data", "template": "https://cdn.tsetmc.com/api/StaticData/GetStaticData", "method": "GET", "description": "Static descriptions/names such as paper type, sectors, YVal"},
    {"category": "StaticData", "name": "Time", "template": "https://cdn.tsetmc.com/api/StaticData/GetTime", "method": "GET", "description": "Server date and time"},
    {"category": "Learning", "name": "Learning topics", "template": "https://cdn.tsetmc.com/api/Learning/GetLearningTopics", "method": "GET", "description": "Learning/help topics from TSETMC site"},
]

DEFAULT_PARAMS = {
    "InsCode": "17914401175772326",   # اهرم
    "DEven": "20260701",
    "Flow": "0",
    "Top": "10",
    "CSecVal": "27",
    "Query": "اهرم",
    "Symbol": "اهرم",
    "TracingNo": "0",
    "MainTableRowId": "1",
    "AuctionId": "1",
    "ShareHolderShareID": "1",
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126 Safari/537.36",
    "Accept": "application/json,text/plain,text/csv,text/html,*/*",
    "Referer": "https://www.tsetmc.com/",
    "Origin": "https://www.tsetmc.com",
}

RETRYABLE_STATUS = {429, 502, 503, 504}
WAF_MARKERS = ["request rejected", "درخواست شما به دلایلی مسدود شده", "access denied", "cloudflare"]

RESULT_FIELDS = [
    "category", "name", "method", "template", "url", "description",
    "status", "status_code", "elapsed_ms", "retry_count", "content_type",
    "bytes", "json_keys", "record_count_hint", "sample", "error", "tested_at",
]


@dataclass
class TestResult:
    category: str = ""
    name: str = ""
    method: str = "GET"
    template: str = ""
    url: str = ""
    description: str = ""
    status: str = ""
    status_code: str = ""
    elapsed_ms: str = ""
    retry_count: str = "0"
    content_type: str = ""
    bytes: str = ""
    json_keys: str = ""
    record_count_hint: str = ""
    sample: str = ""
    error: str = ""
    tested_at: str = ""


def fill_template(template: str, params: Dict[str, str]) -> str:
    url = template
    for k, v in params.items():
        url = url.replace("{" + k + "}", str(v))
    return url


def is_tsetmc_domain(url: str) -> bool:
    return urlparse(url).netloc.lower().endswith("tsetmc.com")


def classify_content(text: str, status_code: int, url: str, content_type: str = "") -> Tuple[str, str, str]:
    stripped = (text or "").strip()
    if status_code >= 400:
        return f"HTTP_{status_code}", "", ""
    if not stripped or stripped in {"[]", "{}"}:
        return "ACTIVE_EMPTY", ("list[0]" if stripped == "[]" else ""), "0"

    # تشخیص فایل Excel باینری (xlsx یک آرشیو ZIP است؛ امضای آن "PK" است)
    ctype_low = (content_type or "").lower()
    is_excel_ctype = any(k in ctype_low for k in ["spreadsheet", "ms-excel", "excel"])
    is_zip_magic = stripped[:2] == "PK"
    if is_excel_ctype or is_zip_magic:
        return "ACTIVE_EXCEL", "", ""  # تعداد ردیف بعداً در test_one با openpyxl محاسبه می‌شود

    low = stripped[:3000].lower()
    if "<!doctype html" in low or "<html" in low:
        if any(m in low for m in WAF_MARKERS):
            return "WAF_BLOCKED", "", ""
        if is_tsetmc_domain(url):
            return "SPA_DEAD", "", ""  # مسیر دیگر وجود ندارد؛ به صفحه اصلی fallback شده
        return "THIRD_PARTY_PAGE", "", ""

    try:
        data = json.loads(stripped)
        json_keys, count_hint = "", ""
        if isinstance(data, dict):
            keys = list(data.keys())
            json_keys = ", ".join(keys[:10])
            for k, v in data.items():
                if isinstance(v, list):
                    count_hint = f"{k}: {len(v)}"
                    break
                if isinstance(v, dict):
                    count_hint = f"{k}: object"
                    break
        elif isinstance(data, list):
            json_keys, count_hint = "list", f"list: {len(data)}"
        return "ACTIVE_JSON", json_keys, count_hint
    except Exception:
        pass

    if "," in stripped[:2000] or ";" in stripped[:2000]:
        return "ACTIVE_CSV_OR_TEXT", "", ""
    return "ACTIVE_TEXT", "", ""


def do_request(session: requests.Session, url: str, timeout: float,
                max_retries: int, base_delay: float) -> Tuple[Optional[requests.Response], str, int]:
    err = ""
    for attempt in range(max_retries + 1):
        try:
            resp = session.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
            if resp.status_code in RETRYABLE_STATUS and attempt < max_retries:
                time.sleep(base_delay * (2 ** attempt))
                continue
            return resp, "", attempt
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
            err = repr(e)
            if attempt < max_retries:
                time.sleep(base_delay * (2 ** attempt))
                continue
            return None, err, attempt
    return None, err, max_retries


def test_one(session: requests.Session, ep: Dict[str, str], params: Dict[str, str],
             timeout: float, max_retries: int, delay: float) -> TestResult:
    url = fill_template(ep["template"], params)
    result = TestResult(
        category=ep["category"], name=ep["name"], method=ep.get("method", "GET"),
        template=ep["template"], url=url, description=ep.get("description", ""),
        tested_at=datetime.now().isoformat(timespec="seconds"),
    )
    t0 = time.perf_counter()
    effective_timeout = float(ep.get("timeout", timeout))
    effective_retries = int(ep.get("max_retries", max_retries))
    resp, err, retries = do_request(session, url, effective_timeout, effective_retries, base_delay=0.6)
    result.elapsed_ms = str(int((time.perf_counter() - t0) * 1000))
    result.retry_count = str(retries)

    if resp is None:
        result.status, result.error = "ERROR", err
    else:
        text = resp.text or ""
        result.status_code = str(resp.status_code)
        result.content_type = resp.headers.get("content-type", "")
        result.bytes = str(len(resp.content or b""))
        status, json_keys, count_hint = classify_content(text, resp.status_code, url, result.content_type)
        result.status = status
        result.json_keys = json_keys
        result.record_count_hint = count_hint

        if status == "ACTIVE_EXCEL" and openpyxl is not None:
            try:
                wb = openpyxl.load_workbook(BytesIO(resp.content), read_only=False, data_only=True)
                ws = wb[wb.sheetnames[0]]
                row_count = ws.max_row
                result.record_count_hint = f"rows: {row_count} (sheet: {wb.sheetnames[0]})"
                result.sample = f"[Excel file — {len(wb.sheetnames)} sheet(s), {row_count} rows]"
            except Exception as e:
                result.record_count_hint = f"excel_parse_error: {e!r}"
                result.sample = "[Excel file — could not parse for row count]"
        else:
            result.sample = text[:250].replace("\r", " ").replace("\n", " ").replace("\t", " ")

    time.sleep(delay)
    return result


def save_csv(results: List[TestResult], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=RESULT_FIELDS)
        w.writeheader()
        for r in results:
            w.writerow({k: getattr(r, k) for k in RESULT_FIELDS})


def save_json(results: List[TestResult], path: Path) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump([asdict(r) for r in results], f, ensure_ascii=False, indent=2)


def print_summary(results: List[TestResult]) -> None:
    counts: Dict[str, int] = {}
    for r in results:
        counts[r.status] = counts.get(r.status, 0) + 1
    print("\nSummary")
    print("=======")
    print(f"Total tested: {len(results)} (baseline: previously confirmed active)")
    for k in sorted(counts):
        print(f"{k:16s} {counts[k]}")
    still_active = sum(1 for r in results if r.status in {"ACTIVE_JSON", "ACTIVE_CSV_OR_TEXT", "ACTIVE_TEXT", "ACTIVE_EXCEL"})
    print(f"\nStill working: {still_active}/{len(results)}")
    broken = [r for r in results if r.status not in {"ACTIVE_JSON", "ACTIVE_CSV_OR_TEXT", "ACTIVE_TEXT", "ACTIVE_EXCEL"}]
    if broken:
        print("\n⚠️ Endpoints that regressed since last test:")
        for r in broken:
            print(f"  - [{r.category}] {r.name} -> {r.status} ({r.status_code or r.error})")


def main() -> None:
    parser = argparse.ArgumentParser(description="Re-test the 55 confirmed-active TSETMC endpoints.")
    parser.add_argument("--inscode", default=DEFAULT_PARAMS["InsCode"])
    parser.add_argument("--date", default=DEFAULT_PARAMS["DEven"], help="YYYYMMDD")
    parser.add_argument("--timeout", type=float, default=20.0)
    parser.add_argument("--delay", type=float, default=0.1)
    parser.add_argument("--max-retries", type=int, default=3)
    parser.add_argument("--csv", default="tsetmc_active_endpoints_test.csv")
    parser.add_argument("--json", default="tsetmc_active_endpoints_test.json")
    args = parser.parse_args()

    params = dict(DEFAULT_PARAMS)
    params["InsCode"] = args.inscode
    params["DEven"] = args.date

    session = requests.Session()
    results: List[TestResult] = []
    for idx, ep in enumerate(ENDPOINTS, 1):
        r = test_one(session, ep, params, args.timeout, args.max_retries, args.delay)
        retry_tag = f" retry={r.retry_count}" if r.retry_count != "0" else ""
        print(f"[{idx:02d}/{len(ENDPOINTS)}] {r.status:16s} {r.status_code or '-':>4s} {r.elapsed_ms:>6s}ms{retry_tag} | {r.category} | {r.name}")
        results.append(r)

    save_csv(results, Path(args.csv))
    save_json(results, Path(args.json))
    print_summary(results)
    print(f"\nSaved: {args.csv}")
    print(f"Saved: {args.json}")


if __name__ == "__main__":
    main()
